#가동분석
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from os import listdir
thin_border=Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),top=Side(border_style="thin", color="000000"),bottom=Side(border_style="thin", color="000000"))#검은색 thin테두리 설정
dnt_border=Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),top=Side(border_style="double", color="000000"),bottom=Side(border_style="thin", color="000000"))
my_font=Font(bold=True,color="FFFFFF") #폰트 설정(볼드,흰색)
fill_B,fill_O=PatternFill("solid",fgColor="4F81BD"),PatternFill("solid",fgColor="F79646") #파란색,오렌지색 채우기
twb=openpyxl.Workbook()
twb.create_sheet("Team_Total")
del twb["Sheet"]
ttws=twb["Team_Total"]
ttws["B1"],ttws["C1"],ttws["D1"],ttws["E1"]="소속/팀","총 인원수","관측자 수","총 관측횟수" #최종시트에 기록
ttws["B1"].fill,ttws["C1"].fill,ttws["D1"].fill,ttws["E1"].fill=fill_B,fill_B,fill_B,fill_B
ttws["B1"].font,ttws["C1"].font,ttws["D1"].font,ttws["E1"].font=my_font,my_font,my_font,my_font
ttws["B1"].border,ttws["B2"].border,ttws["C1"].border,ttws["C2"].border,ttws["D1"].border,ttws["D2"].border,ttws["E1"].border,ttws["E2"].border=thin_border,thin_border,thin_border,thin_border,thin_border,thin_border,thin_border,thin_border
ttws.column_dimensions["B"].width,ttws.column_dimensions["C"].width,ttws.column_dimensions["E"].width,ttws.column_dimensions["E"].width=30,11,12,12
filepath=listdir(".")
fulldict=dict() #최종 팀 종합 act 딕셔너리
for files in filepath:
    if ".xlsx" in str(files): #개별 파일 읽기
        wb=openpyxl.load_workbook(files, data_only=True)
        wblst=sorted([sht for sht in wb.sheetnames if sht.startswith("03") or sht.startswith("04")])
        elst=list(str(files[:-5])) # -5를 해서 파일 확장자(.xlsx)를 뺀다.
        for x in elst:
            if x==" ":
                elst[elst.index(x)]="_"
        files="".join(elst)
        shtname=files[files.index("(")+1:-1] #-8을 해서 ")_result"(8글자)를 뺀다.
        cs=twb.create_sheet(shtname)
        tws=twb[shtname]
        count=0 #개인 총 관측회수
        ws0=wb[wblst[0]]
        tws["B1"],tws["C1"],tws["D1"],tws["E1"]="소속/팀/파트","팀/파트인원","관측자","총 관측횟수"
        tws["B2"],tws["C2"],tws["D2"]=str(ws0["F3"].value)+"/"+str(ws0["H3"].value)+"/"+str(ws0["J3"].value),ws0["L3"].value,ws0["P3"].value
        tws["B1"].fill,tws["C1"].fill,tws["D1"].fill,tws["E1"].fill=fill_O,fill_O,fill_O,fill_O
        tws["B1"].font,tws["C1"].font,tws["D1"].font,tws["E1"].font=my_font,my_font,my_font,my_font
        tws["B1"].border,tws["B2"].border,tws["C1"].border,tws["C2"].border,tws["D1"].border,tws["D2"].border,tws["E1"].border,tws["E2"].border=thin_border,thin_border,thin_border,thin_border,thin_border,thin_border,thin_border,thin_border
        pt=str(ws0["F3"].value)+"/"+str(ws0["H3"].value) #공장/팀만 받아오기 위한 객체
        tws.column_dimensions["B"].width,tws.column_dimensions["C"].width,tws.column_dimensions["E"].width=30,11,12
        actdict=dict() #개인이 바뀔 때 마다 딕셔너리 초기화
        for sht in wblst: #개인 파일의 날짜별 시트 루프를 돈다.
            ws=wb[sht]
            i,l=7,5 #i는 for구문 돌기 위한 변수, l은 count 관련 while구문을 돌기 위한 변수
            while ws.cell(row=6,column=l).value!=None and ws.cell(row=6,column=l).value!="-":
                count+=1
                l+=1
            while ws["B"+str(i)].value!="합 계":
                j,k=5,0
                while ws.cell(row=5, column=j).value!="계": #각 회별 측정값을 더 한다.
                    if ws.cell(row=i,column=j).value==None:
                        ws.cell(row=i,column=j).value="0"
                    k+=int(ws.cell(row=i,column=j).value)
                    j+=1
                if ws["C"+str(i)].value!=None:
                    actdict[ws["C"+str(i)].value]=actdict.get(ws["C"+str(i)].value,0)+k #개별 act딕셔너리 더함.
                    fulldict[ws["C"+str(i)].value]=fulldict.get(ws["C"+str(i)].value,0)+k
                i+=1
        m,total,tws["E2"]=1,0,count #m은 for구문을 위한 변수, total은 개인 총 관측치, E2에 개인 총 관측횟수를 입력
        for key,value in actdict.items():
            tws.cell(row=m+3,column=2).value,tws.cell(row=m+3,column=3).value=key,value
            tws.cell(row=m+3,column=2).border,tws.cell(row=m+3,column=3).border=thin_border,thin_border
            total+=int(tws.cell(row=m+3,column=3).value)
            m+=1
        tws["B"+str(m+3)],tws["C"+str(m+3)],tws["B"+str(m+3)].fill,tws["B"+str(m+3)].font="총합계",total,fill_O,my_font
        for n in range(4,m+4): #원본 m+3이었음
            tws.cell(row=n,column=4).value="=C"+str(n)+"/"+str(total) #개인 개별 관측치/개인 총 관측치
            tws.cell(row=n,column=4).border=thin_border
        tws["B"+str(m+3)].border,tws["C"+str(m+3)].border,tws["D"+str(m+3)].border=dnt_border,dnt_border,dnt_border
twslst=[sht for sht in twb.sheetnames if sht!="Team_Total"]
teamnum,lookernum,tobsvnum=0,0,0 #팀원 수, 관측자 수, 팀 총 관측 횟수
for sht in twslst:
    ws1=twb[sht]
    ttws["B2"]=pt #원본 ws1["B2"].value #맨 마지막 사람 파트를 끌고 오는데 파트를 제거할 방법을 구상할 것.
    teamnum+=int(ws1["C2"].value)
    lookernum+=1
    tobsvnum+=int(ws1["E2"].value)
o,fulltotal=4,0 #o는 for구문을 위한 변수, fulltotal은 팀 통합 관측치
for keys,values in fulldict.items():
    ttws["B"+str(o)],ttws["C"+str(o)]=keys,values
    ttws["B"+str(o)].border,ttws["C"+str(o)].border=thin_border,thin_border
    fulltotal+=int(values)
    o+=1
ttws["B"+str(o)],ttws["B"+str(o)].fill,ttws["B"+str(o)].font,ttws["C"+str(o)],ttws["C2"],ttws["D2"],ttws["E2"]="총합계",fill_B,my_font,fulltotal,teamnum,lookernum,tobsvnum
for p in range(4,o+1): #원본 그냥 o였음
    ttws["D"+str(p)]="=C"+str(p)+"/"+str(fulltotal) #팀 개별 관측치/팀 총 관측치
    ttws["D"+str(p)].border=thin_border
ttws["B"+str(o)].border,ttws["C"+str(o)].border,ttws["D"+str(o)].border=dnt_border,dnt_border,dnt_border
twb.save("Team_Work_Samping_Total.xlsx")
