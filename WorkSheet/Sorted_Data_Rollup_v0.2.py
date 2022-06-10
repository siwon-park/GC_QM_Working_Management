#-*- coding:utf-8 -*-
###sort_data를 돌린 개인별 시트를 파트별로 롤업(팀것도 동일)
import openpyxl, os, operator
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from os import listdir
fill_LightOrange,fill_LightBlue,fill_LightGreen,fill_LightYellow=PatternFill("solid",fgColor="FCE4D6"),PatternFill("solid",fgColor="D9E1F2"),PatternFill("solid",fgColor="E2EFDA"),PatternFill("solid",fgColor="FCFCD6")
my_font,black_bold=Font(bold=True,color="FFFFFF"),Font(bold=True,color="000000")
thin_border=Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),top=Side(border_style="thin", color="000000"),bottom=Side(border_style="thin", color="000000"))
centered_A=Alignment(horizontal='center',vertical='center')
cdir=os.path.abspath('.')
Plantlst=["OCP","HSP","ESP"]
alst=["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
filepath=listdir('.')
for files in filepath:
    if ".xlsx" in files:
        wb=openpyxl.load_workbook(files)
        shtlst=wb.sheetnames
        wb.create_sheet("Part_Total")
        pws=wb["Part_Total"]
        TPWdict=dict()
        for sht in shtlst:
            iws=wb[sht]
            r=3
            while iws["B"+str(r)].value!="일별 합계":
                TPWdict[iws["C"+str(r)].value]=TPWdict.get(iws["C"+str(r)].value,iws["B"+str(r)].value)
                r+=1
        imp=3
        for keys,values in sorted(TPWdict.items(),key=operator.itemgetter(1)): #TPWdict의 내용을 기록
            pws["B"+str(imp)],pws["C"+str(imp)]=values,keys
            imp+=1
        pws["B"+str(imp)],pws["B"+str(imp)].font,pws["B"+str(imp)].fill="일별 합계",black_bold,fill_LightOrange
        for i in range(20):
            Tempdict=dict()
            for sht in shtlst:
                iws=wb[sht]
                r=3
                while iws["B"+str(r)].value!="일별 합계":
                    v=iws.cell(row=r,column=i+6).value
                    if v==None:
                        v=0
                    Tempdict[iws["C"+str(r)].value]=Tempdict.get(iws["C"+str(r)].value,0)+int(v)
                    r+=1
            j=3
            while j<imp:
                for keys,values in Tempdict.items(): #Temp의 내용을 기록
                    try:
                        pws.cell(row=j,column=i+6).value=Tempdict[pws["C"+str(j)].value]
                        pws.cell(row=j,column=i+6).number_format="#,##0"
                    except:
                        continue
                j+=1
        for i in range(2,30): #프로세스, 태스크, 구분, 작업유형, 날짜, 등등 쓰기 + 꾸미기
            pws.cell(row=2,column=i).value,pws.cell(row=2,column=i).font,pws.cell(row=2,column=i).fill=iws.cell(row=2,column=i).value,black_bold,fill_LightOrange
        for i in range(3,imp+1):
            pws["Z"+str(i)].value,pws["Z"+str(i)].number_format="=SUM(F"+str(i)+":Y"+str(i)+")","#,##0" #합계
            pws["AA"+str(i)].value,pws["AA"+str(i)].number_format="=Z"+str(i)+"/Z"+str(imp),"0.0%"#비율
            pws["AB"+str(i)].value,pws["AB"+str(i)].number_format="=IFERROR(AVERAGE(F"+str(i)+":Y"+str(i)+"),0)","0.0"#평균
            pws["AC"+str(i)].value,pws["AC"+str(i)].number_format="=IFERROR(STDEVP(F"+str(i)+":Y"+str(i)+"),0)","0.0"#표준편차
        for i in range(20):#일별 합계 SUM
            pws.cell(row=imp,column=i+6).value,pws.cell(row=imp,column=i+6).number_format="=SUM("+alst[i+5]+"3:"+alst[i+5]+str(imp-1)+")","#,##0"
        for cells in pws["B2:AC"+str(imp)]:
            for cell in cells:
                cell.border,cell.alignment=thin_border,centered_A
        pws.sheet_view.zoomScale=75
        pws.column_dimensions["A"].width,pws.column_dimensions["B"].width,pws.column_dimensions["C"].width,pws.column_dimensions["D"].width,pws.column_dimensions["E"].width,pws.row_dimensions[2].height=2,25,25,15,23,35
        pws["B1"],pws["C1"],pws["D1"],pws["E1"],pws["F1"],pws["H1"],pws["J1"],pws["L1"],pws["N1"],pws["P1"]="소속","QM실","공장",iws["E1"].value,"팀",iws["H1"].value,"파트",iws["L1"].value,"인원수",len(shtlst)
        pws["B1"].font,pws["C1"].font,pws["D1"].font,pws["E1"].font,pws["F1"].font,pws["H1"].font,pws["J1"].font,pws["L1"].font,pws["N1"].font,pws["P1"].font=black_bold,black_bold,black_bold,black_bold,black_bold,black_bold,black_bold,black_bold,black_bold,black_bold
        pws.merge_cells("B"+str(imp)+":E"+str(imp)),pws.merge_cells("F1:G1"),pws.merge_cells("H1:I1"),pws.merge_cells("J1:K1"),pws.merge_cells("L1:M1"),pws.merge_cells("N1:O1"),pws.merge_cells("P1:Q1")
        pws["B1"].fill,pws["D1"].fill,pws["F1"].fill,pws["J1"].fill,pws["N1"].fill=fill_LightBlue,fill_LightBlue,fill_LightBlue,fill_LightBlue,fill_LightBlue
        pws["B1"].alignment=centered_A
        for i in range(2,18):
            pws.cell(row=1,column=i).border=thin_border
        inp=input("팀 or 파트?")
        if inp=="팀":
            rwb=openpyxl.load_workbook("Z:\\QM직무_팀.xlsx",data_only=True)
            teamjobdict=dict()
            for pln in Plantlst: #직무별 주작업 받아오기
                if pln in cdir:
                    rws=rwb[pln]
                    r,c=8,4
                    while rws.cell(row=8,column=c).value!=None:
                        if pws["H1"].value==rws.cell(row=8,column=c).value: #AA1셀에 있는 주작업이 주작업 리스트(행)에 있으면 그 이후로 아래로 내려가서 주작업을 딕셔너리로 받아옴
                            while r+1 <= 57:
                                if rws.cell(row=r+1,column=c).value!=None:
                                    teamjobdict[rws.cell(row=r+1,column=3).value]=teamjobdict.get(rws.cell(row=r+1,column=3).value,"주작업")
                                r+=1
                        c+=1
            for i in range(3,imp):
                pws["E"+str(i)]=teamjobdict.get(pws["B"+str(i)].value)
                if pws["C"+str(i)].value=="휴식" or pws["C"+str(i)].value=="휴가" or pws["C"+str(i)].value=="점심시간":
                    pws["E"+str(i)]="작업여유"
                elif pws["E"+str(i)].value==None:
                    pws["E"+str(i)]="부수작업"
                pws["E"+str(i)].fill=fill_LightYellow
        dv=DataValidation(type="list", formula1='"주작업,부수작업,작업여유"') #반드시 '" 순서로 써줘야함. 엑셀은 "를 텍스트로 인식.
        pws.add_data_validation(dv)
        dv.add("E3:E"+str(imp-1))
        wb.save(files)
