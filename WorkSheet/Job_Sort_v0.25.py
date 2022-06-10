#직무별 일자 피벗팅
import openpyxl,os,operator
from os import listdir
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.chart import BarChart, PieChart, Series, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
fill_LightOrange,fill_LightBlue,fill_LightGreen,fill_LightYellow=PatternFill("solid",fgColor="FCE4D6"),PatternFill("solid",fgColor="D9E1F2"),PatternFill("solid",fgColor="E2EFDA"),PatternFill("solid",fgColor="FCFCD6")
my_font,black_bold=Font(bold=True,color="FFFFFF"),Font(bold=True,color="000000")
thin_border=Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),top=Side(border_style="thin", color="000000"),bottom=Side(border_style="thin", color="000000"))
centered_A=Alignment(horizontal='center',vertical='center')
filepath=listdir(".")
alst=["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
rwb=openpyxl.load_workbook("Z:\\QM직무.xlsx",data_only=True) #"" C:\\Users\\User\\Desktop\\QM직무.xlsx
for files in filepath:
    if ".xlsx" in str(files):
        wb=openpyxl.load_workbook(files)
        shtlst=wb.sheetnames
        for sht in shtlst:
            if "_Total" in sht:
                rsht=sht
                shtlst.remove(rsht)
        jlst=[] #직무를 받아오기 위한 리스트 생성
        for sht in shtlst:
            iws=wb[sht]
            jn=iws["AA1"].value
            if jn not in jlst: #직무를 중복없이 받아옴
                jlst.append(jn)
        for job in jlst:
            wb.create_sheet(job)
            jws,Tjdict,jpnum=wb[job],dict(),0
            for sht in shtlst:
                iws,r=wb[sht],3
                if iws["AA1"].value==job:
                    jpnum+=1
                    while iws["B"+str(r)].value!="일별 합계":
                        Tjdict[iws["C"+str(r)].value]=Tjdict.get(iws["C"+str(r)].value,iws["B"+str(r)].value)
                        r+=1
            imp=3
            for keys,values in sorted(Tjdict.items(),key=operator.itemgetter(1)):
                jws["B"+str(imp)],jws["C"+str(imp)]=values,keys
                imp+=1
            jws["B"+str(imp)],jws["B"+str(imp)].font,jws["B"+str(imp)].fill="일별 합계",black_bold,fill_LightOrange
            jws.merge_cells("B"+str(imp)+":E"+str(imp))
            for i in range(20):
                Tempdict=dict()
                jws.cell(row=imp,column=i+6).value,jws.cell(row=imp,column=i+6).number_format="=SUM("+alst[i+5]+"3:"+alst[i+5]+str(imp-1)+")","#,##0"
                for sht in shtlst:
                    iws,r=wb[sht],3
                    if iws["AA1"].value==job:
                        while iws["B"+str(r)].value!="일별 합계":
                            v=iws.cell(row=r,column=i+6).value
                            if v==None:
                                v=0
                            Tempdict[iws["C"+str(r)].value]=Tempdict.get(iws["C"+str(r)].value,0)+int(v)
                            r+=1
                j=3
                while j<imp:
                    for keys,values in Tempdict.items(): #Temp의 내용을 기록
                        try:
                            jws.cell(row=j,column=i+6).value=Tempdict[jws["C"+str(j)].value]
                            jws.cell(row=j,column=i+6).number_format="#,##0"
                        except:
                            continue
                    j+=1
            for i in range(2,30):
                jws.cell(row=2,column=i).value=iws.cell(row=2,column=i).value
                jws.cell(row=2,column=i).font,jws.cell(row=2,column=i).fill=black_bold,fill_LightOrange
            for i in range(3,imp+1):
                jws["Z"+str(i)].value,jws["Z"+str(i)].number_format="=SUM(F"+str(i)+":Y"+str(i)+")","#,##0" #합계
                jws["AA"+str(i)].value,jws["AA"+str(i)].number_format="=Z"+str(i)+"/Z"+str(imp),"0.0%"#비율
                jws["AB"+str(i)].value,jws["AB"+str(i)].number_format="=IFERROR(AVERAGE(F"+str(i)+":Y"+str(i)+"),0)","0.0"#평균
                jws["AC"+str(i)].value,jws["AC"+str(i)].number_format="=IFERROR(STDEVP(F"+str(i)+":Y"+str(i)+"),0)","0.0"#표준편차
            jws.sheet_view.zoomScale=75
            jws.column_dimensions["A"].width,jws.column_dimensions["B"].width,jws.column_dimensions["C"].width,jws.column_dimensions["D"].width,jws.column_dimensions["E"].width,jws.row_dimensions[2].height=2,25,28,11,20,35
            jws["B1"],jws["C1"],jws["D1"],jws["E1"],jws["F1"],jws["H1"],jws["J1"],jws["L1"],jws["O1"],jws["Q1"]="소속","QM실","공장",iws["E1"].value,"팀",iws["H1"].value,"직무",job,"인원수",jpnum
            jws["B1"].fill,jws["D1"].fill,jws["F1"].fill,jws["J1"].fill,jws["O1"].fill=fill_LightBlue,fill_LightBlue,fill_LightBlue,fill_LightBlue,fill_LightBlue
            jws.merge_cells("F1:G1"),jws.merge_cells("H1:I1"),jws.merge_cells("J1:K1"),jws.merge_cells("L1:N1"),jws.merge_cells("O1:P1"),jws.merge_cells("Q1:R1")
            for i in range(2,19):
                jws.cell(row=1,column=i).font,jws.cell(row=1,column=i).border,jws.cell(row=1,column=i).alignment=black_bold,thin_border,centered_A
            for cells in jws["B2:AC"+str(imp)]:
                for cell in cells:
                    cell.border,cell.alignment=thin_border,centered_A
            rws=rwb[jws["E1"].value]
            r,c,teamjobdict=8,4,dict()
            while rws.cell(row=8,column=c).value!=None: ##여기 고치기
                if jws["L1"].value==rws.cell(row=8,column=c).value: #L1셀에 있는 직무가 주작업 리스트(행)에 있으면 그 이후로 아래로 내려가서 주작업을 딕셔너리로 받아옴
                    while r+1 <= 57:
                        if rws.cell(row=r+1,column=c).value!=None:
                            teamjobdict[rws.cell(row=r+1,column=3).value]=teamjobdict.get(rws.cell(row=r+1,column=3).value,"주작업")
                        r+=1
                c+=1
            for i in range(3,imp):
                jws["E"+str(i)]=teamjobdict.get(jws["B"+str(i)].value)
                if jws["C"+str(i)].value=="휴식" or jws["C"+str(i)].value=="휴가" or jws["C"+str(i)].value=="점심시간":
                    jws["E"+str(i)]="작업여유"
                elif jws["E"+str(i)].value==None:
                    jws["E"+str(i)]="부수작업"
                jws["E"+str(i)].fill=fill_LightYellow
            dv=DataValidation(type="list", formula1='"주작업,부수작업,작업여유"') #반드시 '" 순서로 써줘야함. 엑셀은 "를 텍스트로 인식.
            jws.add_data_validation(dv)
            dv.add("E3:E"+str(imp-1))
            jws["B"+str(imp+3)],jws["C"+str(imp+3)],jws["D"+str(imp+3)]="항목","항목별 시간(분)","항목별 비율"
            jws["B"+str(imp+4)],jws["B"+str(imp+5)],jws["B"+str(imp+6)],jws["B"+str(imp+7)]="주작업","부수작업","작업여유","소계"
            jws["C"+str(imp+4)],jws["C"+str(imp+5)],jws["C"+str(imp+6)],jws["C"+str(imp+7)]="=SUMIF(E3:E"+str(imp-1)+',"주작업",Z3:Z'+str(imp-1)+")","=SUMIF(E3:E"+str(imp-1)+',"부수작업",Z3:Z'+str(imp-1)+")","=SUMIF(E3:E"+str(imp-1)+',"작업여유",Z3:Z'+str(imp-1)+")","=SUM(C"+str(imp+4)+":C"+str(imp+6)+")"
            jws["C"+str(imp+4)].number_format,jws["C"+str(imp+5)].number_format,jws["C"+str(imp+6)].number_format,jws["C"+str(imp+7)].number_format="#,##0","#,##0","#,##0","#,##0"
            jws["D"+str(imp+4)],jws["D"+str(imp+5)],jws["D"+str(imp+6)],jws["D"+str(imp+7)]="=C"+str(imp+4)+"/C"+str(imp+7),"=C"+str(imp+5)+"/C"+str(imp+7),"=C"+str(imp+6)+"/C"+str(imp+7),"=C"+str(imp+7)+"/C"+str(imp+7)
            jws["B"+str(imp+3)].fill,jws["C"+str(imp+3)].fill,jws["D"+str(imp+3)].fill,jws["B"+str(imp+4)].fill,jws["B"+str(imp+5)].fill,jws["B"+str(imp+6)].fill,jws["B"+str(imp+7)].fill=fill_LightGreen,fill_LightGreen,fill_LightGreen,fill_LightGreen,fill_LightGreen,fill_LightGreen,fill_LightGreen
            for i in range(3,8):
                for j in range(2,5):
                    jws.cell(row=imp+3,column=j).font=black_bold
                    jws.cell(row=imp+i,column=j).border=thin_border
            for i in range(4,8):
                jws["D"+str(imp+i)].number_format="0.0%"
            chart1,chart2=BarChart(),PieChart() #바 그래프, 파이 그래프
            chart1.type="col" #열 기준
            chart1.style=3 # 스타일 3(스타일 바꾸면 색만 바뀜)
            chart1.title,chart2.title="항목별 시간","항목별 비율" #차트 제목
            cate1=Reference(jws,min_col=2,min_row=imp+4,max_row=imp+6) #카테고리(=범주, 축의 데이터)
            data1,data2=Reference(jws,min_col=3,min_row=imp+3,max_row=imp+6),Reference(jws,min_col=4,min_row=imp+4,max_row=imp+6) #차트 내용 데이터
            chart1.add_data(data1, titles_from_data=True),chart2.add_data(data2) #차트에 데이터 입력(데이터로 부터 데이터 계열 제목 받아옴)
            chart1.set_categories(cate1),chart2.set_categories(cate1) #차트에 축 설정
            #데이터 레이블 설정(데이터 레이블 표시)
            s1,s2=chart1.series[0],chart2.series[0]
            s1.dLbls,s2.dLbls=DataLabelList(),DataLabelList()
            s1.dLbls.showVal,s2.dLbls.showVal=True,True
            chart1.legend.position,chart2.legend.position="b","b"
            jws.add_chart(chart1, "F"+str(imp+3)), jws.add_chart(chart2, "O"+str(imp+3))
            inp=input("팀?")
            if inp=="팀" or inp=="xla":
                jpnum=0
                for sht in shtlst:
                    iws=wb[sht]
                    if iws["L1"].value==job or iws["AA1"].value==job:
                        jpnum+=int(iws["Q1"].value)
                jws["Q1"]=jpnum
        wb.save(files)
