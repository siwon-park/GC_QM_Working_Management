#-*- coding:utf-8 -*-
#정상근무/한외근무 날짜별 피벗팅 프로그램 #시트 맨 뒤에 정상근무, 한외근무 2개 만듦.
import openpyxl,os
import operator #딕셔너리 정렬을 위해 오퍼레이터 import
from os import listdir
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
def AssignName(filename): #이름 할당 함수
    name=""
    for ltr in filename[::-1][13:]:
        if ltr.isalpha() is not True:
            break
        else:
            name+=ltr
    return name[::-1]
my_font,black_bold=Font(bold=True,color="FFFFFF"),Font(bold=True,color="000000")
thin_border=Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),top=Side(border_style="thin", color="000000"),bottom=Side(border_style="thin", color="000000"))#검은색 thin테두리 설정
cdir=os.path.abspath('..')
prsdir=cdir[::-1][:cdir[::-1].index("\\")][::-1]
teamlist=["QA(ESP)","QC(ESP)","QA(HSP)","QC(HSP)","GQM","QA(compliance)","QA(OCP)","QC(OCP)","QE(OCP)","QM지원"]
filepath=listdir(".")
alst=["F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]#크기조절을 위한 알파벳 리스트(len:21)
daylst=["03/16","03/17","03/18","03/19","03/20","03/23","03/24","03/25","03/26","03/27","03/30","03/31","04/01","04/02","04/03","04/06","04/07","04/08","04/09","04/10","소계"] #len:21
for files in filepath:
    if ")_result" in str(files):
        print("현재 읽고있는 중인 파일:",str(files))
        wb=openpyxl.load_workbook(files, data_only=True)
        ws=wb["Total"]
        wb.create_sheet("정상근무"),wb.create_sheet("한외근무")
        fws,ews=wb["정상근무"],wb["한외근무"]
        fws["B2"],fws["C2"],fws["D2"],fws["E2"]="프로세스(L3)","태스크(L4)","구분","작업유형(주/부/여유)"
        ews["B2"],ews["C2"],ews["D2"],ews["E2"]="프로세스(L3)","태스크(L4)","구분","작업유형(주/부/여유)"
        fws["B2"].font,fws["C2"].font,fws["D2"].font,fws["E2"].font=black_bold,black_bold,black_bold,black_bold
        ews["B2"].font,ews["C2"].font,ews["D2"].font,ews["E2"].font=black_bold,black_bold,black_bold,black_bold
        fws.column_dimensions["B"].width,fws.column_dimensions["C"].width,fws.column_dimensions["D"].width,fws.column_dimensions["E"].width,fws.row_dimensions[2].height=25,25,15,15,35
        ews.column_dimensions["B"].width,ews.column_dimensions["C"].width,ews.column_dimensions["D"].width,ews.column_dimensions["E"].width,ews.row_dimensions[2].height=25,25,15,15,35
        for i in range(21):
            fws.cell(row=2,column=i+6).value,fws.cell(row=2,column=i+6).font=daylst[i],black_bold
            fws.column_dimensions[alst[i]].width=6
            ews.cell(row=2,column=i+6).value,ews.cell(row=2,column=i+6).font=daylst[i],black_bold
            ews.column_dimensions[alst[i]].width=6
        fws.column_dimensions["Z"].width,ews.column_dimensions["Z"].width=10,10
        Tdict=dict() #모든 태스크(키)와 프로세스(밸류)를 받아오기 위한 딕셔너리
        i=0 #Tdict를 위한 변수
        while ws.cell(row=i+3,column=4).value!=None:
            Tdict[ws.cell(row=i+3,column=4).value]=Tdict.get(ws.cell(row=i+3,column=4).value,ws.cell(row=i+3,column=3).value)
            i+=1
        imp=3 #Tdict값을 할당하기 위한 변수(imp)
        for keys,values in sorted(Tdict.items(),key=operator.itemgetter(1)): #(value를 기준으로 정렬한)Tdict의 키와 값을 fws에 기록
            fws["B"+str(imp)],fws["C"+str(imp)]=values,keys
            ews["B"+str(imp)],ews["C"+str(imp)]=values,keys
            imp+=1 #imp는 나중에 써야함 혹시 모르니 일반적으론 쓸때 -1할 것(최종적으로 +1된 상태이니)
        j,k=3,6
        for i in range(len(daylst)):
            tempNdict,tempEXdict=dict(),dict() #일자별 임시 딕셔너리,일자별 한외 딕셔너리 생성
            while ws["A"+str(j)].value==daylst[i]:
                tempNdict[ws.cell(row=j,column=4).value]=tempNdict.get(ws.cell(row=j,column=4).value,0)+int(ws.cell(row=j,column=9).value)
                tempEXdict[ws.cell(row=j,column=4).value]=tempEXdict.get(ws.cell(row=j,column=4).value,0)+int(ws.cell(row=j,column=10).value)
                j+=1
            for n in range(3,imp):
                try:
                    fws.cell(row=n,column=k).value=tempNdict[fws["C"+str(n)].value]
                    ews.cell(row=n,column=k).value=tempEXdict[ews["C"+str(n)].value]
                except:
                    continue
            k+=1
        fws["B"+str(imp)],fws["B"+str(imp)].font="일별 합계",black_bold
        ews["B"+str(imp)],ews["B"+str(imp)].font="일별 합계",black_bold
        fws.merge_cells("B"+str(imp)+":"+"E"+str(imp)) #Bimp에서 Eimp까지 셀 합병
        ews.merge_cells("B"+str(imp)+":"+"E"+str(imp))
        for i in range(21):#일자별 합계 SUM함수 적용
            fws.cell(row=imp,column=i+6).value="=SUM("+alst[i]+"3:"+alst[i]+str(imp-1)+")"
            ews.cell(row=imp,column=i+6).value="=SUM("+alst[i]+"3:"+alst[i]+str(imp-1)+")"
        for i in range(3,imp):#태스크별 합계 SUM함수 적용
            fws.cell(row=i,column=26).value="=SUM(F"+str(i)+":Y"+str(i)+")"
            ews.cell(row=i,column=26).value="=SUM(F"+str(i)+":Y"+str(i)+")"
            fws["B"+str(i)].value,fws["C"+str(i)].value=fws["B"+str(i)].value[fws["B"+str(i)].value.index("/")+2:],fws["C"+str(i)].value[fws["C"+str(i)].value.index("/")+2:] #프로세스, 태스크 코드 없애기
            ews["B"+str(i)].value,ews["C"+str(i)].value=ews["B"+str(i)].value[ews["B"+str(i)].value.index("/")+2:],ews["C"+str(i)].value[ews["C"+str(i)].value.index("/")+2:]
        for cells in fws["B2:Z"+str(imp)]: #모든범위에 테두리 칠하기
            for cell in cells:
                cell.border=thin_border
                if str(cell.value)=="0":
                    cell.value=None
        for cells in ews["B2:Z"+str(imp)]: #모든범위에 테두리 칠하기
            for cell in cells:
                cell.border=thin_border
                if str(cell.value)=="0":
                    cell.value=None
        fws.sheet_view.zoomScale,ews.sheet_view.zoomScale=80,80 #워크시트 확대/축소 비율 80%로 바꾸기.
        fws.insert_rows(1),ews.insert_rows(1)
        wb.save(str(files)) #있는 파일에 그대로 덮어쓰기
#프로세스, 태스크 정렬은 코드를 기준으로 정렬했다가 나중에 코드를 삭제한 채로 오름차순 정렬된 것임!
#필드값 없음이 있으면 읽어오지 못 한다. 보고 나중에 값이 없으면 0으로 바꾸거나 해야할듯
