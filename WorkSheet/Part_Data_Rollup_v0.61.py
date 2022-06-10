###파트데이터 롤업###
import openpyxl
import os
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from os import listdir
cdir=os.path.abspath('..')
prsdir=cdir[::-1][:cdir[::-1].index("\\")][::-1] #파트/팀명(상위폴더 기준으로)
my_font,black_bold=Font(bold=True,color="FFFFFF"),Font(bold=True,color="000000") #폰트 설정(볼드,흰색)
fill_B=PatternFill("solid",fgColor="4F81BD") #파란색 채우기
fill_GC=PatternFill("solid",fgColor="EF7A54") #GC색 옅게 셀 채우기
thin_border=Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),top=Side(border_style="thin", color="000000"),bottom=Side(border_style="thin", color="000000"))#검은색 thin테두리 설정
under_line=Border(bottom=Side(border_style="thin", color="000000")) #밑에만 테두리
teamlist=["QA(ESP)","QC(ESP)","QA(HSP)","QC(HSP)","GQM","QA(compliance)","QA(OCP)","QC(OCP)","QE(OCP)","QM지원"]
filepath=listdir('.')
twb=openpyxl.Workbook() #참고할 자료 엑셀 파일이 열려있으면 안 만들어짐. 오류남 Access denied.
twt=twb.create_sheet("Part_Total")
ttws=twb["Part_Total"] #twb.get_sheet_by_name("Part_Total")
for lst in teamlist: #팀명을 B1셀에 쓰기 위함.
    if lst in cdir:
        ttws["B1"]=lst
if "(OCP)" in str(ttws["B1"].value) or "(ESP)" in str(ttws["B1"].value) or "(HSP)" in str(ttws["B1"].value):
    ttws["B1"]=str(ttws["B1"].value)[:2]
ttws["A1"],ttws["C1"],ttws["D1"]="팀명","파트명",prsdir
ttws["A2"],ttws["B2"],ttws["C2"],ttws["D2"],ttws["E2"],ttws["F2"]="공장","팀","파트","직무","이름","Glevel"
ttws["G2"],ttws["H2"],ttws["I2"],ttws["J2"],ttws["K2"],ttws["L2"]="날짜","작업분류","프로세스(L3)","태스크(L4)","SOP No. 또는 Product code","SOP 명칭 또는 제품명"
ttws["M2"],ttws["N2"],ttws["O2"],ttws["P2"]="시험/검토/승인(수)","이론시간(분)","실제시간(분)","기타(한외근무)"
for files in filepath:
    if ".xlsx" in str(files):
        wb=openpyxl.load_workbook(files, data_only=True)
        ws=wb["Total"] #wb.get_sheet_by_name("Total")
        elst=list(str(files[:-5]))# -5를 해서 파일 확장자(.xlsx)를 뺀다.
        for x in elst:
            if x==" ":
                elst[elst.index(x)]="_"
        files="".join(elst)
        shtname=files[files.index("(")+1:-8] #-8을 해서 ")_result"(8글자)를 뺀다.
        pws=twb.create_sheet(shtname)
        tws=twb[shtname] #twb.get_sheet_by_name(shtname)
        i=1
        while True: #개인별 시트의 Total 옮기기
            cv1,cv2,cv3,cv4,cv5,cv6,cv7,cv8,cv9,cv10=ws["A"+str(i)].value,ws["B"+str(i)].value,ws["C"+str(i)].value,ws["D"+str(i)].value,ws["E"+str(i)].value,ws["F"+str(i)].value,ws["G"+str(i)].value,ws["H"+str(i)].value,ws["I"+str(i)].value,ws["J"+str(i)].value
            tws["A"+str(i)],tws["B"+str(i)],tws["C"+str(i)]=cv1,cv2,cv3
            tws["D"+str(i)],tws["E"+str(i)],tws["F"+str(i)]=cv4,cv5,cv6
            tws["G"+str(i)],tws["H"+str(i)],tws["I"+str(i)],tws["J"+str(i)]=cv7,cv8,cv9,cv10
            if ws["A"+str(i)].value != ws["A"+str(i+1)].value: #날짜가 바뀔 때 밑줄 긋기
                for row in tws["A"+str(i):"J"+str(i)]:
                    for cell in row:
                        cell.border=under_line
            i+=1
            lst=[cv1,cv2,cv3,cv4,cv5,cv6,cv7,cv8,cv9,cv10]
            if all([cvs==None for cvs in lst]):
                tws.column_dimensions["A"].width,tws.column_dimensions["B"].width,tws.column_dimensions["C"].width=7,11,15
                tws.column_dimensions["D"].width,tws.column_dimensions["E"].width,tws.column_dimensions["F"].width=15,28,22
                tws.column_dimensions["G"].width,tws.column_dimensions["H"].width,tws.column_dimensions["I"].width,tws.column_dimensions["I"].width=19,13,13,15
                for m in range(10): #통합 변경(셀 폰트, 셀 채우기, 셀 테두리)
                    tws.cell(row=2,column=m+1).font=my_font
                    tws.cell(row=2,column=m+1).fill=fill_GC
                    tws.cell(row=2,column=m+1).border=thin_border
                tws["A1"].font,tws["C1"].font,tws["E1"].font,tws["G1"].font=my_font,my_font,my_font,my_font #특정 셀 폰트 변경
                tws["A1"].fill,tws["C1"].fill,tws["E1"].fill,tws["G1"].fill=fill_GC,fill_GC,fill_GC,fill_GC #특정 셀 색 채우기
                tws["A1"].border,tws["C1"].border,tws["E1"].border,tws["G1"].border=thin_border,thin_border,thin_border,thin_border #(특정 셀) 테두리 그리기
                tws.auto_filter.ref = "A2:J2"
                break
pn=""
if "OCP" in str(cdir):
    pn="OCP"
elif "ESP" in str(cdir):
    pn="ESP"
elif "HSP" in str(cdir):
    pn="HSP"
elst=twb.sheetnames #twb.get_sheet_names()
elst.remove("Part_Total")
elst.remove("Sheet")
j,k=3,1
for shtnames in elst:
    ws=twb[shtnames] #twb.get_sheet_by_name(shtnames)
    while True:
        tcv1,tcv2,tcv3,tcv4,tcv5,tcv6,tcv7,tcv8,tcv9,tcv10=ws["A"+str(k+2)].value,ws["B"+str(k+2)].value,ws["C"+str(k+2)].value,ws["D"+str(k+2)].value,ws["E"+str(k+2)].value,ws["F"+str(k+2)].value,ws["G"+str(k+2)].value,ws["H"+str(k+2)].value,ws["I"+str(k+2)].value,ws["J"+str(k+2)].value
        ttws["A"+str(j)],ttws["B"+str(j)],ttws["C"+str(j)],ttws["D"+str(j)],ttws["E"+str(j)],ttws["F"+str(j)]=pn,ttws["B1"].value,ttws["D1"].value,ws["D1"].value,ws["F1"].value,ws["H1"].value
        ttws["G"+str(j)],ttws["H"+str(j)],ttws["I"+str(j)]=tcv1,tcv2,tcv3
        ttws["J"+str(j)],ttws["K"+str(j)],ttws["L"+str(j)]=tcv4,tcv5,tcv6
        ttws["M"+str(j)],ttws["N"+str(j)],ttws["O"+str(j)],ttws["P"+str(j)]=tcv7,tcv8,tcv9,tcv10
        j+=1
        k+=1
        lst2=[tcv1,tcv2,tcv3,tcv4,tcv5,tcv6,tcv7,tcv8,tcv9,tcv10]
        if all([tcvs==None for tcvs in lst2]):
            ttws["A"+str(j-1)],ttws["B"+str(j-1)],ttws["C"+str(j-1)],ttws["D"+str(j-1)],ttws["E"+str(j-1)],ttws["F"+str(j-1)]=None,None,None,None,None,None #맨마지막 빈행에 입력되는 데이터 제거
            k=1
            j-=1
            break
del twb["Sheet"]
ttws.column_dimensions["A"].width,ttws.column_dimensions["B"],ttws.column_dimensions["C"],ttws.column_dimensions["D"],ttws.column_dimensions["E"],ttws.column_dimensions["F"]
ttws.column_dimensions["G"].width,ttws.column_dimensions["H"].width,ttws.column_dimensions["I"].width,ttws.column_dimensions["J"].width,ttws.column_dimensions["K"].width,ttws.column_dimensions["L"].width=16,11,11,15,15,28
ttws.column_dimensions["M"].width,ttws.column_dimensions["N"].width,ttws.column_dimensions["O"].width,ttws.column_dimensions["J"].width,ttws.column_dimensions["P"].width=22,19,13,13,15
for l in range(16): #통합 변경(셀 폰트, 셀 채우기, 셀 테두리)
    ttws.cell(row=2,column=l+1).font=my_font
    ttws.cell(row=2,column=l+1).fill=fill_B
    ttws.cell(row=2,column=l+1).border=thin_border
ttws["A1"].font,ttws["A1"].fill,ttws["A1"].border=my_font,fill_B,thin_border
ttws["C1"].font,ttws["C1"].fill,ttws["C1"].border=my_font,fill_B,thin_border
ttws["B1"].font,ttws["D1"].font=black_bold,black_bold
ttws.auto_filter.ref = "A2:P2"
twb.save(prsdir+"_Part_Total_result.xlsx")
