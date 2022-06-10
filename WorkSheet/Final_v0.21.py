#-*- coding:utf-8 -*-
import openpyxl,os,operator
from os import listdir
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
filepath=listdir(".")
fill_LightRed,fill_Purple=PatternFill("solid",fgColor="F2DCDB"),PatternFill("solid",fgColor="CC99FF")
fill_M,fill_S,fill_R=PatternFill("solid",fgColor="CCFF99"),PatternFill("solid",fgColor="FFDB69"),PatternFill("solid",fgColor="CCECFF")
my_font,black_bold=Font(bold=True,color="FFFFFF"),Font(bold=True,color="000000")
thin_border=Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),top=Side(border_style="thin", color="000000"),bottom=Side(border_style="thin", color="000000"))
centered_A=Alignment(horizontal='center',vertical='center')
L3lst=["신제품기술이전관리","QC업무교육훈련및자격부여","시험법밸리데이션_AMV","QC자재구매","초자류관리","시약관리",
"표준품관리","비교품관리","시액조제관리","동물시험관리","검체채취관리","재공시험및시험관리","원자재시험관리",
"안정성시험관리","특별시험관리","위수탁시험관리","기준일탈관리","LIMS시스템운영","보관품관리","변경관리","교육관리",
"직무승인","업무관리","GMP문서관리","공급업체관리","출하관리","일탈관리","고객불만관리","자율점검","실사관리",
"회수관리","반품관리","부적합품관리","Data경향분석","연간품질검토_APQR","허가관리","품질정책관리","위험관리",
"GMP위원회","QMS업무지시","품질기획","Data_Integrity","반제품사용승인","밸리데이션관리","엑셀시트밸리데이션",
"컴퓨터시스템밸리데이션","설비사용관리","교정관리","컴퓨터시스템관리","휴가","휴식","정기미팅","상시미팅","행사(월례조회 등)","사내교육","사외교육","점심시간","메일확인","시험실 청소"," 일일 점검",
"행정업무","미팅 준비","업무 보고","현장 점검","문서 영문화","프로젝트업무"]
alst=["H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ",
"BA","BB","BC","BD","BE","BF","BG","BH","BI","BJ","BK","BL","BM","BN","BO","BP","BQ","BR","BS","BT","BU"] #,"BV","BW","BX","BY","BZ"
nwb=openpyxl.Workbook()
nwb.create_sheet("Total")
del nwb["Sheet"]
tws=nwb["Total"]
tws["A2"],tws["B2"],tws["C2"],tws["D2"],tws["E2"],tws["F2"],tws["G2"]="공장","팀","파트","이름","성별","레벨","직무"
tws["A2"].font,tws["B2"].font,tws["C2"].font,tws["D2"].font,tws["E2"].font,tws["F2"].font,tws["G2"].font=black_bold,black_bold,black_bold,black_bold,black_bold,black_bold,black_bold
tws["A2"].fill,tws["B2"].fill,tws["C2"].fill,tws["D2"].fill,tws["E2"].fill,tws["F2"].fill,tws["G2"].fill=fill_LightRed,fill_LightRed,fill_LightRed,fill_LightRed,fill_LightRed,fill_LightRed,fill_LightRed
r=3 #개인별 공장,팀,파트,이름,레벨,성별,직무 등을 받아오기 위한 r, 3행부터 출발함.
for files in filepath:
    if ".xlsx" in files:
        print("현재 읽고 있는 파일: "+str(files))
        wb=openpyxl.load_workbook(files, data_only=True)
        namelst=wb.sheetnames #워크북을 열었을 때 이름만 있는 리스트
        i=0
        for L3 in L3lst: #2행에 L3리스트 쓰기
            tws.cell(row=2,column=i+8).value,tws.cell(row=2,column=i+8).font,tws.cell(row=2,column=i+8).fill=L3,black_bold,fill_LightRed
            i+=1
        for name in namelst: #개인별 공장,팀,파트,이름,레벨,성별,직무 등을 받아오기
            nws=wb[name] #각 개인별 이름이 써있는 시트를 열고 정보입력
            ndict,row,jdict,M,S,R=dict(),3,dict(),0,0,0
            tws["A"+str(r)],tws["B"+str(r)],tws["C"+str(r)],tws["D"+str(r)],tws["E"+str(r)],tws["F"+str(r)],tws["G"+str(r)]=nws["E1"].value,nws["H1"].value,nws["L1"].value,nws["P1"].value,nws["T1"].value,nws["X1"].value,nws["AA1"].value
            while nws["B"+str(row)].value!="일별 합계":
                ndict[nws["B"+str(row)].value]=ndict.get(nws["B"+str(row)].value,0)+int(nws["Z"+str(row)].value)
                jdict[nws["B"+str(row)].value]=jdict.get(nws["B"+str(row)].value,nws["E"+str(row)].value)
                row+=1
            for i in range(66):
                tws.cell(row=r,column=i+8).value=ndict.get(tws.cell(row=2,column=i+8).value)
                tws.cell(row=r,column=i+8).number_format="#,##0"
                if jdict.get(tws.cell(row=2,column=i+8).value)=="주작업" and tws.cell(row=r,column=i+8).value!=0:
                    tws.cell(row=r,column=i+8).fill=fill_M
                    M+=int(tws.cell(row=r,column=i+8).value)
                elif jdict.get(tws.cell(row=2,column=i+8).value)=="부수작업" and tws.cell(row=r,column=i+8).value!=0:
                    tws.cell(row=r,column=i+8).fill=fill_S
                    S+=int(tws.cell(row=r,column=i+8).value)
                elif jdict.get(tws.cell(row=2,column=i+8).value)=="작업여유" and tws.cell(row=r,column=i+8).value!=0:
                    tws.cell(row=r,column=i+8).fill=fill_R
                    R+=int(tws.cell(row=r,column=i+8).value)
            tws["BW"+str(r)],tws["BW"+str(r)].number_format="=IFERROR("+str(M)+"/BV"+str(r)+",0)","0.0%" #주작업비율
            tws["BX"+str(r)],tws["BX"+str(r)].number_format="=IFERROR("+str(S)+"/BV"+str(r)+",0)","0.0%" #부수작업비율
            tws["BY"+str(r)],tws["BY"+str(r)].number_format="=IFERROR("+str(R)+"/BV"+str(r)+",0)","0.0%" #작업여유비율
            r+=1 #마지막엔 r이 이제 최종 데이터 값이 행+1임
        tws["A1"],tws["B1"],tws["C1"],tws["D1"],tws["D1"].font,tws["D1"].fill="주작업","부수작업","작업여유","합계(L3)",my_font,fill_Purple
        tws["A1"].font,tws["B1"].font,tws["C1"].font,tws["A1"].fill,tws["B1"].fill,tws["C1"].fill=black_bold,black_bold,black_bold,fill_M,fill_S,fill_R
        tws["BV2"],tws["BW2"],tws["BX2"],tws["BY2"]="합계","주작업비율","부수작업비율","작업여유비율"
        tws["BV2"].font,tws["BW2"].font,tws["BX2"].font,tws["BY2"].font=black_bold,black_bold,black_bold,black_bold
        tws["BV2"].fill,tws["BW2"].fill,tws["BX2"].fill,tws["BY2"].fill=fill_Purple,fill_M,fill_S,fill_R
        for row in range(3,r):
            tws["BV"+str(row)],tws["BV"+str(row)].number_format="=SUM(H"+str(row)+":BU"+str(row)+")","#,##0"
        tws.merge_cells("D1:G1")
        for i in range(66):
            tws[alst[i]+"1"],tws[alst[i]+"1"].number_format="=SUBTOTAL(9,"+alst[i]+"3:"+alst[i]+str(r-1)+")","#,##0"
        tws.auto_filter.ref,tws.column_dimensions["G"].width="A2:BY2",14
        for cells in tws["A1:BY"+str(r-1)]:
            for cell in cells:
                cell.border,cell.alignment=thin_border,centered_A
                if cell.value==0:
                    cell.value=None
        nwb.save("All_Total.xlsx")
